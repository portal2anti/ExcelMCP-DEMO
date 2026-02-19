#!/usr/bin/env python3
"""
Enrich testData.xlsx with a 'Capacity' scenario: month-level unused capacity
(people, FTE) per account. Adds columns Unused Headcount, Unused FTE and new
Capacity scenario rows.
"""
import random
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

OUTPUT_FILE = Path(__file__).resolve().parent / "testData.xlsx"
SHEET_NAME = "testData"

# Same dimensions as generator for new Capacity rows
ENTITIES = ["Entity A", "Entity B", "Entity C", "Entity D"]
REGIONS = ["North America", "EMEA", "APAC", "LATAM"]
DEPARTMENTS = ["Sales", "Marketing", "R&D", "G&A", "Operations"]
PRODUCT_LINES = ["Product Alpha", "Product Beta", "Product Gamma", "Product Delta", "Services"]

def main():
    wb = load_workbook(OUTPUT_FILE)
    ws = wb[SHEET_NAME]

    # Insert two columns after Comment (column 20) -> new cols 21, 22
    ws.insert_cols(21, 2)
    ws.cell(row=1, column=21, value="Unused Headcount")
    ws.cell(row=1, column=22, value="Unused FTE")

    # Fill existing data rows with empty capacity (they are actuals/budget/forecast/prior year)
    max_row = ws.max_row
    for r in range(2, max_row + 1):
        ws.cell(row=r, column=21, value="")
        ws.cell(row=r, column=22, value="")

    # Append Capacity scenario rows: one per (Year, Period, Entity, Region, Department, Product)
    # Each row represents unused capacity (people, FTE) at month level; "account" is which cost pool
    next_row = max_row + 1
    years = [2023, 2024]
    periods = list(range(1, 13))

    for year in years:
        for period in periods:
            for entity in ENTITIES:
                for region in REGIONS:
                    for dept in DEPARTMENTS:
                        for product in PRODUCT_LINES:
                            period_end = date(year, min(period, 12), 28 if period != 2 else 28)
                            # Unused capacity in people and FTE (month level)
                            unused_hc = random.randint(0, 25)
                            unused_fte = round(random.uniform(0, 12.5), 2)

                            ws.cell(row=next_row, column=1, value=entity)
                            ws.cell(row=next_row, column=2, value=region)
                            ws.cell(row=next_row, column=3, value=dept)
                            ws.cell(row=next_row, column=4, value=product)
                            ws.cell(row=next_row, column=5, value="Capacity")
                            ws.cell(row=next_row, column=6, value="Capacity")
                            ws.cell(row=next_row, column=7, value=year)
                            ws.cell(row=next_row, column=8, value=period)
                            ws.cell(row=next_row, column=9, value=period_end)
                            ws.cell(row=next_row, column=10, value=0)
                            ws.cell(row=next_row, column=11, value=0)
                            ws.cell(row=next_row, column=12, value=0)
                            ws.cell(row=next_row, column=13, value=0)
                            ws.cell(row=next_row, column=14, value=0)
                            ws.cell(row=next_row, column=15, value=0)
                            ws.cell(row=next_row, column=16, value=0)
                            ws.cell(row=next_row, column=17, value=0)
                            ws.cell(row=next_row, column=18, value=0)
                            ws.cell(row=next_row, column=19, value=0)
                            ws.cell(row=next_row, column=20, value="Unused capacity (month)")
                            ws.cell(row=next_row, column=21, value=unused_hc)
                            ws.cell(row=next_row, column=22, value=unused_fte)
                            next_row += 1

    wb.save(OUTPUT_FILE)
    capacity_rows = next_row - 1 - max_row
    print(f"Enriched {OUTPUT_FILE}: added columns 'Unused Headcount', 'Unused FTE'; "
          f"added {capacity_rows} 'Capacity' scenario rows (month-level unused people/FTE). "
          f"Total rows: {next_row - 1}.")


if __name__ == "__main__":
    main()
