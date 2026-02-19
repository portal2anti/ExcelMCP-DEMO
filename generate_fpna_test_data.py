#!/usr/bin/env python3
"""
Generate a large denormalized Excel table 'testData' with FP&A-style test data.
Single sheet, many rows, typical financial planning columns.
"""
import random
from datetime import date
from pathlib import Path

from openpyxl import Workbook

# Output path in project
OUTPUT_FILE = Path(__file__).resolve().parent / "testData.xlsx"
SHEET_NAME = "testData"

# FP&A-style dimensions
ENTITIES = ["Entity A", "Entity B", "Entity C", "Entity D"]
REGIONS = ["North America", "EMEA", "APAC", "LATAM"]
DEPARTMENTS = ["Sales", "Marketing", "R&D", "G&A", "Operations"]
PRODUCT_LINES = ["Product Alpha", "Product Beta", "Product Gamma", "Product Delta", "Services"]
SCENARIOS = ["Actual", "Budget", "Forecast", "Prior Year", "Capacity"]
ACCOUNT_TYPES = ["Revenue", "COGS", "OpEx", "CapEx", "Headcount", "Other Income", "Depreciation"]

def main():
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    # Header row (denormalized: entity, period, scenario, region, dept, product, account type, amounts, etc.)
    headers = [
        "Entity",
        "Region",
        "Department",
        "Product Line",
        "Account Type",
        "Scenario",
        "Fiscal Year",
        "Period",
        "Period End Date",
        "Revenue",
        "COGS",
        "Gross Profit",
        "OpEx",
        "EBITDA",
        "Depreciation",
        "EBIT",
        "Headcount",
        "CapEx",
        "Other Income",
        "Comment",
        "Unused Headcount",
        "Unused FTE",
    ]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h)

    row = 2
    years = [2023, 2024]  # 2 years × 12 periods × 4×4×5×5×4 = 38,400 rows
    periods = list(range(1, 13))  # 1-12

    # One denormalized row per (year, period, entity, region, department, product, scenario)
    for year in years:
        for period in periods:
            for entity in ENTITIES:
                for region in REGIONS:
                    for dept in DEPARTMENTS:
                        for product in PRODUCT_LINES:
                            for scenario in SCENARIOS:
                                period_end = date(year, min(period, 12), 28 if period != 2 else 28)
                                is_capacity = scenario == "Capacity"
                                if is_capacity:
                                    rev = cogs = gross = opex = ebitda = depr = ebit = 0
                                    hc = 0
                                    capex = other = 0
                                    comment = "Unused capacity (month)"
                                    unused_hc = random.randint(0, 25)
                                    unused_fte = round(random.uniform(0, 12.5), 2)
                                    acc_type = "Capacity"
                                else:
                                    rev = round(random.uniform(500, 15000), 2)
                                    cogs = round(random.uniform(200, 6000), 2)
                                    gross = round(rev - cogs, 2)
                                    opex = round(random.uniform(100, 3000), 2)
                                    ebitda = round(gross - opex, 2)
                                    depr = round(random.uniform(20, 400), 2)
                                    ebit = round(ebitda - depr, 2)
                                    hc = random.randint(5, 120)
                                    capex = round(random.uniform(0, 800), 2)
                                    other = round(random.uniform(-200, 200), 2)
                                    comment = random.choice(["", "", "", "Q4 push", "One-time", "Reclass", "Accrual", "FX"])
                                    unused_hc = ""
                                    unused_fte = ""
                                    acc_type = random.choice(ACCOUNT_TYPES)

                                ws.cell(row=row, column=1, value=entity)
                                ws.cell(row=row, column=2, value=region)
                                ws.cell(row=row, column=3, value=dept)
                                ws.cell(row=row, column=4, value=product)
                                ws.cell(row=row, column=5, value=acc_type)
                                ws.cell(row=row, column=6, value=scenario)
                                ws.cell(row=row, column=7, value=year)
                                ws.cell(row=row, column=8, value=period)
                                ws.cell(row=row, column=9, value=period_end)
                                ws.cell(row=row, column=10, value=rev)
                                ws.cell(row=row, column=11, value=cogs)
                                ws.cell(row=row, column=12, value=gross)
                                ws.cell(row=row, column=13, value=opex)
                                ws.cell(row=row, column=14, value=ebitda)
                                ws.cell(row=row, column=15, value=depr)
                                ws.cell(row=row, column=16, value=ebit)
                                ws.cell(row=row, column=17, value=hc)
                                ws.cell(row=row, column=18, value=capex)
                                ws.cell(row=row, column=19, value=other)
                                ws.cell(row=row, column=20, value=comment)
                                ws.cell(row=row, column=21, value=unused_hc)
                                ws.cell(row=row, column=22, value=unused_fte)
                                row += 1

    # Freeze header
    ws.freeze_panes = "A2"
    # Auto-fit would need openpyxl.utils or manual width; leave default

    wb.save(OUTPUT_FILE)
    print(f"Saved {OUTPUT_FILE} with sheet '{SHEET_NAME}', rows 1-{row - 1} ({row - 2} data rows).")

if __name__ == "__main__":
    main()
